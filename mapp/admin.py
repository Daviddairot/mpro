from django.contrib import admin
from django.contrib import messages
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from .models import Grade, Student, Assessment, CA, DeletedStudent

admin.site.site_header = "Musicology Admin"
admin.site.site_title = "Musicology Admin"
admin.site.index_title = "Musicology Admin"

@admin.action(description='Export selected assessments to Excel')
def export_assessments_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Assessments"

    headers = [
        'S/N', 'Assessor', 'First Name', 'Last Name', 'Matric Number', 'Instrument',
        'Song 1', 'Song 2', 'Song 3', 'Dressing', 'Total'
    ]
    ws.append(headers)

    for idx, assessment in enumerate(queryset.select_related('student', 'assessor'), start=1):
        ws.append([
            idx,
            assessment.assessor.username if assessment.assessor else "N/A",
            assessment.student.first_name,
            assessment.student.last_name,
            assessment.student.matric_number,
            assessment.student.instrument,
            float(assessment.song1),
            float(assessment.song2),
            float(assessment.song3),
            float(assessment.dressing),
            float(assessment.total)
        ])

    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, len(header) + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="assessments_export.xlsx"'
    wb.save(response)
    return response

@admin.action(description='Export selected grades to Excel')
def export_grades_to_excel(modeladmin, request, queryset):
    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"

    headers = [
        'S/N', 'Student', 'Matric Number', 'CBT', 'Practical', 'Classwork',
        'Assignment', 'Total CA', 'Exam', 'Total'
    ]
    ws.append(headers)

    for idx, grade in enumerate(queryset.select_related('student', 'ca'), start=1):
        ws.append([
            idx,
            f"{grade.student.first_name} {grade.student.last_name}",
            grade.student.matric_number,
            float(grade.ca.CBT) if grade.ca else 0,
            float(grade.ca.practical) if grade.ca else 0,
            float(grade.ca.classwork) if grade.ca else 0,
            float(grade.ca.Assignment) if grade.ca else 0,
            float(grade.ca.total) if grade.ca else 0,
            float(grade.score),
            float(grade.total)
        ])

    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, len(header) + 2)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="grades_export.xlsx"'
    wb.save(response)
    return response


class SoftDeleteAdminMixin:
    """
    Deleting from the admin (single object or 'Delete selected') sends rows to the
    recycle bin instead of removing them, since Django's bulk queryset.delete()
    normally bypasses each model's overridden delete(). Single-object deletes
    already go through delete_model() -> obj.delete(), which is soft by default.
    """
    def delete_queryset(self, request, queryset):
        for obj in queryset:
            obj.delete()


class StudentAdmin(SoftDeleteAdminMixin, admin.ModelAdmin):
    list_display = ('first_name', 'last_name', 'matric_number', 'instrument')
    search_fields = ('first_name', 'last_name', 'matric_number', 'instrument')

class AssessmentAdmin(SoftDeleteAdminMixin, admin.ModelAdmin):
    list_display = ('student', 'song1', 'song2', 'song3', 'dressing', 'assessor', 'total')
    search_fields = ('student__matric_number', 'student__first_name', 'student__last_name')
    list_filter = ('assessor',)
    actions = [export_assessments_to_excel]

class CAAdmin(SoftDeleteAdminMixin, admin.ModelAdmin):
    list_display = ('student', 'CBT', 'practical', 'classwork', 'Assignment', 'total', 'assessor')
    search_fields = ('student__matric_number', 'student__first_name', 'student__last_name')

@admin.register(Grade)
class GradeAdmin(SoftDeleteAdminMixin, admin.ModelAdmin):
    list_display = ('student', 'score', 'ca', 'extra', 'total')
    search_fields = ('student__first_name', 'student__last_name', 'student__matric_number')
    list_filter = ('student__instrument',)
    readonly_fields = ('score', 'total')
    autocomplete_fields = ('student', 'ca')
    actions = [export_grades_to_excel]

admin.site.register(Student, StudentAdmin)
admin.site.register(Assessment, AssessmentAdmin)
admin.site.register(CA, CAAdmin)


@admin.action(description='Restore selected students')
def restore_students(modeladmin, request, queryset):
    count = queryset.count()
    for student in queryset:
        student.restore()
    modeladmin.message_user(request, f"Restored {count} student(s).", messages.SUCCESS)


@admin.action(description='Permanently delete selected students (cannot be undone)')
def hard_delete_students(modeladmin, request, queryset):
    count = queryset.count()
    for student in queryset:
        student.hard_delete()
    modeladmin.message_user(request, f"Permanently deleted {count} student(s).", messages.WARNING)


@admin.register(DeletedStudent)
class DeletedStudentAdmin(admin.ModelAdmin):
    list_display = ('first_name', 'last_name', 'matric_number', 'instrument', 'deleted_at')
    search_fields = ('first_name', 'last_name', 'matric_number', 'instrument')
    ordering = ('-deleted_at',)
    actions = [restore_students, hard_delete_students]

    def get_queryset(self, request):
        return Student.all_objects.filter(is_deleted=True)

    def get_actions(self, request):
        actions = super().get_actions(request)
        actions.pop('delete_selected', None)
        return actions

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False
