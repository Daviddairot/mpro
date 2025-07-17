# musicapp/views.py
from django.http import HttpResponse
from openpyxl import load_workbook
from mapp.utils import get_tokens_for_user
from .models import Grade, Student, Assessment
from .forms import AssessmentForm
from django.contrib.auth import login, authenticate
from django.contrib.auth.forms import UserCreationForm
from django.shortcuts import get_object_or_404, render, redirect
from django.contrib.auth.decorators import login_required
from .models import Student, Assessment, CA
from .forms import AssessmentForm, StudentSearchForm
import logging
from docx import Document
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from django.contrib.auth.models import User
from rest_framework import status
from django.shortcuts import get_object_or_404
from rest_framework.decorators import api_view, parser_classes
from rest_framework.parsers import MultiPartParser, FormParser
from django.views.decorators.csrf import csrf_exempt 
import traceback
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


logger = logging.getLogger(__name__)


@api_view(['POST'])
@permission_classes([AllowAny])
def signup(request):
    username = request.data.get('username')
    password = request.data.get('password')
    if User.objects.filter(username=username).exists():
        return Response({'error': 'Username already exists'}, status=400)
    user = User.objects.create_user(username=username, password=password)
    tokens = get_tokens_for_user(user)
    return Response({'message': 'User created successfully', 'tokens': tokens})


@api_view(['POST'])
@permission_classes([AllowAny])
def login_view(request):
    username = request.data.get("username")
    password = request.data.get("password")

    if not username or not password:
        return Response({"error": "Username and password are required"}, status=status.HTTP_400_BAD_REQUEST)

    user = authenticate(username=username, password=password)
    
    if not user:
        return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)
    
    tokens = get_tokens_for_user(user)

    students = Student.objects.all()
    
    studentData = [
        {
            'id': student.id,
            'firstName': student.first_name or "NULL",
            'lastName': student.last_name or "NULL",
            'matricNumber': student.matric_number or "NULL",
            'instrument': student.instrument or "NULL"
        }
        for student in students
    ]
    return Response({'message': 'Login successful', 'tokens': tokens, 'students': studentData})


@api_view(['POST'])
@permission_classes([IsAuthenticated])
def ca(request):
    user = request.user
    student_id = request.data.get('student')  # Assuming student is passed as an ID
    CBT = request.data.get('CBT')
    practical = request.data.get('practical')
    classwork = request.data.get('classwork')
    Assignments = request.data.get('Assignments')
    print(student_id)
    # Get Assessor (User)
    try:
        assessor = User.objects.get(username=user.username)
    except User.DoesNotExist:
        return Response({'error': 'Invalid Assessor'}, status=status.HTTP_400_BAD_REQUEST)

    # Get Student
    student = get_object_or_404(Student, id=student_id)
    print(student)
    # Get or create CA record
    ca, created = CA.objects.get_or_create(student=student)
    print(ca)
    # Assign assessor
    ca.assessor = assessor
    

    # Update only if a value is provided
    if CBT is not None:
        ca.CBT = CBT
    if practical is not None:
        ca.practical = practical
    if classwork is not None:
        ca.classwork = classwork
    if Assignments is not None:
        ca.Assignment = Assignments  # Make sure field name matches your model

    ca.save()

    return Response({'message': 'Success'}, status=status.HTTP_200_OK)

@api_view(['GET'])
@permission_classes([IsAuthenticated])
def getCA(request, id):
    user = request.user
    stud = get_object_or_404(Student, id=id)
    student, created = CA.objects.get_or_create(student=stud)
    studentData = [
        {
            # 'id': student.id,
            'firstName': student.student.first_name or "NULL",
            'lastName': student.student.last_name or "NULL",
            'matricNumber': student.student.matric_number or "NULL",
            'instrument': student.student.instrument or "NULL",
            "CBT": student.CBT,
            "practical": student.practical,
            "classwork": student.classwork,
            "Assignment": student.Assignment,
            "total": student.total
        }
    ]
    return Response({'ca':studentData}, status=200)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def getUser(request):
    print("got here")
    user = request.user
    return Response (
        {'user': user.username}, 
        status = 201)

@api_view(['POST'])
@permission_classes([IsAuthenticated])
def exam(request):
    user = request.user
    student_id = request.data.get('student')
    song1 = request.data.get('song1')
    song2 = request.data.get('song2')
    song3 = request.data.get('song3')
    dressing = request.data.get('dressing')
    
    # Get student
    student = get_object_or_404(Student, id=student_id)

    # Create assessment
    examData = {
        'song1': song1,
        'song2': song2,
        'song3': song3,
        'dressing': dressing,
        'student': student,
        'assessor': user,
    }

    assessment = Assessment.objects.create(**examData)
    
    return Response({'message': 'success'}, status=201)



def grades_list(request):
    grades = Grade.objects.all()
    return Response({'grades': grades}, status=201)


def download_grades(request):

    wb = Workbook()
    ws = wb.active
    ws.title = "Grades"

    # Header row
    headers = [
        'S/N', 'Student', 'Matric Number', 'Exam', 'CBT', 'Practical',
        'Classwork', 'Assignment', 'Total CA', 'Total'
    ]
    ws.append(headers)

    # Data rows
    for idx, grade in enumerate(Grade.objects.all(), start=1):
        ws.append([
            idx,
            f"{grade.student.first_name} {grade.student.last_name}",
            grade.student.matric_number,
            grade.score,
            grade.ca.CBT,
            grade.ca.practical,
            grade.ca.classwork,
            grade.ca.Assignment,
            grade.ca.total,
            # grade.extra,
            grade.total
        ])

    # Optional: set column widths
    for i, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = max(10, len(header) + 2)

    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=grades.xlsx'
    wb.save(response)
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def studentData(request):
    user = request.user
    assessor =  get_object_or_404(User, username=user)
    if assessor is None:
        return Response({'error': 'Invalid Assessor'}, status=status.HTTP_400_BAD_REQUEST)
    students = Student.objects.all()
    studentData = [
        {
            'id': student.id,
            'firstName': student.first_name or "NULL",
            'lastName': student.last_name or "NULL",
            'matricNumber': student.matric_number or "NULL",
            'instrument': student.instrument or "NULL"
        }
        for student in students
    ]
    return Response({'message': 'all registered student', 'students': studentData})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
def grade(request):
    user = request.user
    assessor =  get_object_or_404(User, username=user)
    if assessor is None:
        return Response({'error': 'Invalid Assessor'}, status=status.HTTP_400_BAD_REQUEST)
    grades = Grade.objects.all()
    gradeData = [
        {
            'id': grade.id,
            'extra': grade.extra,
            'firstname': grade.student.first_name,
            'lastname':grade.student.last_name,
            'matric_number': grade.student.matric_number,
            'CA': grade.ca.total,
            'Exam': grade.score,
            'total': grade.total
        }
        for grade in grades
    ]
    return Response({'message': 'all grades', 'grades': gradeData})


@api_view(['POST'])
@csrf_exempt
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def import_students_from_docx(request):
    uploaded_file = request.FILES.get('file')

    if not uploaded_file:
        return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        document = Document(uploaded_file)
        table = document.tables[0]
        created_count = 0
        updated_count = 0

        for row in table.rows[1:]:  # Skip header
            matric_number = row.cells[0].text.strip()
            full_name = row.cells[1].text.strip()
            instrument = row.cells[2].text.strip()

            surname, first_name = parse_name(full_name)

            student, created = Student.objects.update_or_create(
                matric_number=matric_number,
                defaults={
                    'first_name': first_name,
                    'last_name': surname,
                    'instrument': instrument
                }
            )
            if created:
                created_count += 1
            else:
                updated_count += 1

        return Response({
            "message": "Import completed successfully.",
            "created": created_count,
            "updated": updated_count
        })

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

def parse_name(full_name):
    parts = full_name.split(", ")
    return (parts[0], parts[1]) if len(parts) == 2 else ('', full_name)


# Utility to read Excel rows
def parse_excel(uploaded_file):
    workbook = load_workbook(uploaded_file)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    return rows[1:]  # Skip header

@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def import_classwork_scores(request):
    uploaded_file = request.FILES.get("file")
    if not uploaded_file:
        return Response({"error": "No file uploaded."}, status=400)

    try:
        rows = parse_excel(uploaded_file)
        updated = 0
        not_found = []

        for row in rows:
            try:
                matric, score = str(row[0]).strip(), row[1]
                student = Student.objects.get(matric_number__iexact=matric)
                ca_obj, _ = CA.objects.get_or_create(student=student)
                ca_obj.classwork = float(score)
                ca_obj.save()
                updated += 1
            except Student.DoesNotExist:
                not_found.append(matric)
            except Exception as e:
                print("Row error:", e)
                continue

        return Response({
            "message": "CBT import complete",
            "updated": updated,
            "not_found": not_found
        })
    except Exception as e:
        traceback.print_exc()
        return Response({"error": str(e)}, status=500)


@api_view(['POST'])
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def import_practical_scores(request):
    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = parse_excel(uploaded_file)
        updated = 0
        not_found = []

        for row in rows:
            matric_number, score_text = str(row[0]).strip(), row[1]
            try:
                student = Student.objects.get(matric_number__iexact=matric_number)
                ca_obj, _ = CA.objects.get_or_create(student=student)
                score = float(score_text)
                ca_obj.practical = score
                ca_obj.save()
                updated += 1
            except (ValueError, TypeError):
                continue
            except Student.DoesNotExist:
                not_found.append(matric_number)

        return Response({
            "message": "Practical scores import completed.",
            "updated": updated,
            "not_found": not_found
        })

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)


@api_view(['POST'])
@csrf_exempt
@permission_classes([AllowAny])
@parser_classes([MultiPartParser, FormParser])
def import_cbt_scores(request):
    uploaded_file = request.FILES.get('file')
    if not uploaded_file:
        return Response({"error": "No file uploaded."}, status=status.HTTP_400_BAD_REQUEST)

    try:
        rows = parse_excel(uploaded_file)
        updated = 0
        not_found = []
        invalid_scores = []

        for row in rows:
            matric_number, score_text = str(row[0]).strip(), row[1]
            try:
                score = float(score_text)
            except (ValueError, TypeError):
                invalid_scores.append(matric_number)
                continue

            try:
                student = Student.objects.get(matric_number__iexact=matric_number)
                ca_obj, _ = CA.objects.get_or_create(student=student)
                ca_obj.CBT = score
                ca_obj.save()
                updated += 1
            except Student.DoesNotExist:
                not_found.append(matric_number)

        return Response({
            "message": "CBT scores import completed.",
            "updated": updated,
            "not_found": not_found,
            "invalid_scores": invalid_scores,
        })

    except Exception as e:
        return Response({"error": str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
